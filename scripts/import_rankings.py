from __future__ import annotations

import argparse
import csv
import sqlite3
from collections import defaultdict
from pathlib import Path
from typing import Any
from uuid import UUID, uuid5

from openpyxl import load_workbook

from app.database import migrate, read_connection, transaction

NAMESPACE = UUID("c1b85144-1b07-4bd8-96c2-75616a8f8c9c")
REQUIRED_COLUMNS = {"総合順位", "メニュー名", "カテゴリ"}


def read_rows(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook["ランキング"]
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value) for value in next(iterator)]
        return [dict(zip(headers, row, strict=True)) for row in iterator if row[0] is not None]
    with path.open(encoding="utf-8-sig", newline="") as stream:
        return list(csv.DictReader(stream))


def validate(rows: list[dict[str, Any]], expected_count: int | None) -> list[dict[str, Any]]:
    if not rows or not REQUIRED_COLUMNS.issubset(rows[0]):
        raise ValueError(f"required columns: {', '.join(sorted(REQUIRED_COLUMNS))}")
    normalized = [
        {
            "rank": int(row["総合順位"]),
            "name": str(row["メニュー名"]).strip(),
            "category": str(row["カテゴリ"]).strip(),
        }
        for row in rows
    ]
    if expected_count is not None and len(normalized) != expected_count:
        raise ValueError(f"expected {expected_count} rows, got {len(normalized)}")
    ranks = [row["rank"] for row in normalized]
    names = [row["name"] for row in normalized]
    if len(ranks) != len(set(ranks)) or ranks != list(range(1, len(ranks) + 1)):
        raise ValueError("ranks must be unique and contiguous from 1")
    if len(names) != len(set(names)):
        raise ValueError("menu names must be unique")
    if any(not row["category"] for row in normalized):
        raise ValueError("category must not be empty")
    return normalized


def apply_rows(
    database_path: Path,
    rows: list[dict[str, Any]],
    guessed: list[str],
    replace: bool,
) -> None:
    migrate(database_path)
    if database_path.exists() and database_path.stat().st_size:
        backup_dir = database_path.parent / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / f"{database_path.stem}.before-import.sqlite3"
        with sqlite3.connect(database_path) as source, sqlite3.connect(backup_path) as target:
            source.backup(target)
    with transaction(database_path, immediate=True) as connection:
        count = connection.execute("SELECT COUNT(*) FROM menus").fetchone()[0]
        if count and not replace:
            raise ValueError(
                "database already contains menus; "
                "pass --replace after reviewing the backup"
            )
        if replace:
            connection.execute("DELETE FROM guesses")
            connection.execute("DELETE FROM menus")
            connection.execute("DELETE FROM categories")

        categories: dict[str, int] = {}
        category_names = dict.fromkeys(row["category"] for row in rows)
        for category_order, category_name in enumerate(category_names):
            cursor = connection.execute(
                "INSERT INTO categories(name, display_order) VALUES (?, ?)",
                (category_name, category_order),
            )
            categories[category_name] = int(cursor.lastrowid)

        by_category: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            by_category[row["category"]].append(row)
        guessed_set = set(guessed)
        known_names = {row["name"] for row in rows}
        unknown_guesses = guessed_set - known_names
        if unknown_guesses:
            raise ValueError(f"guessed menus not found: {', '.join(sorted(unknown_guesses))}")

        for category_name, category_rows in by_category.items():
            sorted_rows = sorted(category_rows, key=lambda item: item["name"])
            for display_order, row in enumerate(sorted_rows):
                menu_id = str(uuid5(NAMESPACE, row["name"]))
                connection.execute(
                    """
                    INSERT INTO menus(id, name, category_id, rank, display_order)
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (menu_id, row["name"], categories[category_name], row["rank"], display_order),
                )
                if row["name"] in guessed_set:
                    connection.execute("INSERT INTO guesses(menu_id) VALUES (?)", (menu_id,))


def main() -> None:
    parser = argparse.ArgumentParser(description="Import private ranking data into SQLite")
    parser.add_argument("source", type=Path)
    parser.add_argument("--db", type=Path, required=True)
    parser.add_argument("--expected-count", type=int)
    parser.add_argument("--guessed", action="append", default=[])
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--replace", action="store_true")
    args = parser.parse_args()

    rows = validate(read_rows(args.source), args.expected_count)
    print(f"validated {len(rows)} menus; ranks 1-{len(rows)}")
    print(f"categories: {len(set(row['category'] for row in rows))}")
    print(f"initially answered: {len(args.guessed)}")
    if not args.apply:
        print("dry-run only; pass --apply to write the database")
        return
    apply_rows(args.db, rows, args.guessed, args.replace)
    with read_connection(args.db) as connection:
        print(f"imported: {connection.execute('SELECT COUNT(*) FROM menus').fetchone()[0]}")


if __name__ == "__main__":
    main()
